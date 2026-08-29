from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from beadsketch.engine import PatternResult
from beadsketch.excel_io import export_pattern_xlsx, import_pattern_xlsx
from beadsketch.palettes import load_palette


def sample_result() -> PatternResult:
    palette = load_palette("MARD 291")[:8]
    indices = np.arange(30, dtype=np.int32).reshape(5, 6) % len(palette)
    rgb = np.asarray([c.rgb for c in palette], dtype=np.uint8)[indices]
    return PatternResult(
        indices, palette, np.arange(len(palette)), rgb.copy(), rgb.copy(),
        np.zeros((5, 6), dtype=np.float32), "测试", {"brand_palette": "MARD 291"},
    )


class ExcelRoundTripTests(unittest.TestCase):
    def test_untouched_workbook_is_lossless(self) -> None:
        result = sample_result()
        with tempfile.TemporaryDirectory() as folder:
            path = export_pattern_xlsx(result, Path(folder) / "pattern.xlsx")
            restored, report = import_pattern_xlsx(path)
            np.testing.assert_array_equal(restored.indices, result.indices)
            self.assertEqual(report.fill_edits, 0)
            self.assertEqual(report.code_edits, 0)
            self.assertGreaterEqual(len(restored.palette), 291)

    def test_each_used_color_gets_a_location_sheet_and_swatch(self) -> None:
        result = sample_result()
        with tempfile.TemporaryDirectory() as folder:
            path = export_pattern_xlsx(result, Path(folder) / "locations.xlsx")
            wb = load_workbook(path)
            used = result.counts()
            used_codes = [color.code for color, _count in used]
            self.assertEqual(wb.sheetnames[:1 + len(used_codes)], ["拼豆图纸", *used_codes])

            color, count = used[0]
            color_index = next(i for i, item in enumerate(result.palette) if item.code == color.code)
            ws = wb[color.code]
            black_cells = 0
            for y in range(result.height):
                for x in range(result.width):
                    cell = ws.cell(y + 2, x + 2)
                    if int(result.indices[y, x]) == color_index:
                        black_cells += 1
                        self.assertEqual(cell.value, color.code)
                        self.assertEqual(cell.fill.fgColor.rgb[-6:], "000000")
                        self.assertEqual(cell.font.color.rgb[-6:], "FFFFFF")
                    else:
                        self.assertIsNone(cell.value)
            self.assertEqual(black_cells, count)

            swatch_col = result.width + 4
            for row in range(2, 5):
                for col in range(swatch_col, swatch_col + 3):
                    self.assertEqual(ws.cell(row, col).fill.fgColor.rgb[-6:], "%02X%02X%02X" % color.rgb)
            self.assertEqual(ws.cell(3, swatch_col + 1).value, color.code)

    def test_whole_row_fill_wins_over_unchanged_codes(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = export_pattern_xlsx(sample_result(), Path(folder) / "row_gray.xlsx")
            wb = load_workbook(path)
            ws = wb["拼豆图纸"]
            for col in range(2, 8):
                ws.cell(4, col).fill = PatternFill("solid", fgColor="808080")
            wb.save(path)
            restored, report = import_pattern_xlsx(path)
            self.assertEqual(report.fill_edits, 6)
            self.assertEqual(report.code_edits, 0)
            self.assertEqual(len(set(int(v) for v in restored.indices[2])), 1)

    def test_code_only_edit_and_conflict_rule(self) -> None:
        result = sample_result()
        with tempfile.TemporaryDirectory() as folder:
            path = export_pattern_xlsx(result, Path(folder) / "edits.xlsx")
            wb = load_workbook(path)
            ws = wb["拼豆图纸"]
            ws["B2"] = result.palette[4].code
            ws["C2"] = result.palette[5].code
            ws["C2"].fill = PatternFill("solid", fgColor="808080")
            wb.save(path)
            restored, report = import_pattern_xlsx(path)
            self.assertEqual(int(restored.indices[0, 0]), 4)
            self.assertEqual(report.code_edits, 2)
            self.assertEqual(report.fill_edits, 1)
            self.assertEqual(report.conflicts, 1)


if __name__ == "__main__":
    unittest.main()
