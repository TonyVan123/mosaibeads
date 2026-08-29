from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

from .color import pairwise_delta_e, rgb_to_lab
from .engine import PatternResult
from .palettes import BeadColor, load_palette


FORMAT_MARKER = "MOSAIBeads Excel Pattern"
FORMAT_VERSION = 2
PATTERN_SHEET = "拼豆图纸"
PALETTE_SHEET = "完整色板"
DATA_SHEET = "MOSAIBeads_Data"
GRID_START_ROW = 2
GRID_START_COL = 2


@dataclass
class ExcelImportReport:
    source_path: Path
    brand: str
    width: int
    height: int
    fill_edits: int = 0
    code_edits: int = 0
    approximate_matches: int = 0
    conflicts: int = 0
    unchanged: int = 0

    @property
    def message(self) -> str:
        return (
            f"{self.width}×{self.height} · {self.brand} · "
            f"填充色修改 {self.fill_edits} 格 · 色号修改 {self.code_edits} 格 · "
            f"近似匹配 {self.approximate_matches} 格 · 冲突 {self.conflicts} 格"
        )


def _hex(rgb: tuple[int, int, int]) -> str:
    return "%02X%02X%02X" % rgb


def _text_hex(rgb: tuple[int, int, int]) -> str:
    lum = 0.2126 * rgb[0] + 0.7152 * rgb[1] + 0.0722 * rgb[2]
    return "151922" if lum > 150 else "FFFFFF"


def _resolve_brand(result: PatternResult) -> str:
    brand = str(result.metadata.get("brand_palette") or result.metadata.get("palette") or "MARD 291")
    try:
        load_palette(brand)
    except Exception:
        brand = "MARD 291"
    return brand


def _complete_palette(result: PatternResult, brand: str) -> list[BeadColor]:
    colors = list(result.palette)
    seen = {c.code.upper() for c in colors}
    try:
        for color in load_palette(brand):
            if color.code.upper() not in seen:
                colors.append(color)
                seen.add(color.code.upper())
    except Exception:
        pass
    return colors


def export_pattern_xlsx(result: PatternResult, path: str | Path) -> Path:
    """Write visual cell fills plus lossless brand/code metadata."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    brand = _resolve_brand(result)
    full_palette = _complete_palette(result, brand)
    wb = Workbook()
    ws = wb.active
    ws.title = PATTERN_SHEET
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "B2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    header_fill = PatternFill("solid", fgColor="253247")
    header_font = Font(name="Microsoft YaHei UI", size=9, bold=True, color="FFFFFF")
    fine = Side(style="thin", color="7B8491")
    medium = Side(style="medium", color="414B5A")
    board = Side(style="medium", color="D94A4A")
    ws.cell(1, 1, "行/列")
    ws.cell(1, 1).fill, ws.cell(1, 1).font = header_fill, header_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    for x in range(result.width):
        cell = ws.cell(1, x + GRID_START_COL, x + 1)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(x + GRID_START_COL)].width = 4.2
    ws.column_dimensions["A"].width = 6.2
    ws.row_dimensions[1].height = 22

    for y in range(result.height):
        label = ws.cell(y + GRID_START_ROW, 1, y + 1)
        label.fill, label.font = header_fill, header_font
        label.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[y + GRID_START_ROW].height = 24
        for x in range(result.width):
            color = result.palette[int(result.indices[y, x])]
            cell = ws.cell(y + GRID_START_ROW, x + GRID_START_COL, color.code)
            cell.fill = PatternFill("solid", fgColor=_hex(color.rgb))
            cell.font = Font(name="Arial", size=8, bold=True, color=_text_hex(color.rgb))
            cell.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
            left = board if x % 29 == 0 else (medium if x % 5 == 0 else fine)
            top = board if y % 29 == 0 else (medium if y % 5 == 0 else fine)
            right = board if (x + 1) % 29 == 0 or x == result.width - 1 else fine
            bottom = board if (y + 1) % 29 == 0 or y == result.height - 1 else fine
            cell.border = Border(left=left, top=top, right=right, bottom=bottom)
    ws.auto_filter.ref = f"A1:{get_column_letter(result.width + 1)}{result.height + 1}"
    ws.print_area = f"A1:{get_column_letter(result.width + 1)}{result.height + 1}"

    palette_ws = wb.create_sheet(PALETTE_SHEET)
    palette_ws.sheet_view.showGridLines = False
    palette_ws.freeze_panes = "A2"
    palette_ws.append(["色号", "名称", "屏幕 RGB", "色块", "说明"])
    for cell in palette_ws[1]:
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    used = {c.code.upper() for c, _count in result.counts()}
    for color in full_palette:
        palette_ws.append([color.code, color.name, "#" + _hex(color.rgb), "", "图纸已使用" if color.code.upper() in used else "可选"])
        row = palette_ws.max_row
        palette_ws.cell(row, 4).fill = PatternFill("solid", fgColor=_hex(color.rgb))
        palette_ws.cell(row, 4).border = Border(left=fine, right=fine, top=fine, bottom=fine)
    for col, width in {"A": 13, "B": 22, "C": 15, "D": 12, "E": 14}.items():
        palette_ws.column_dimensions[col].width = width
    palette_ws.auto_filter.ref = f"A1:E{palette_ws.max_row}"

    data_ws = wb.create_sheet(DATA_SHEET)
    data_ws["A1"] = FORMAT_MARKER
    for key, value, row in (
        ("format_version", FORMAT_VERSION, 2), ("brand", brand, 3),
        ("width", result.width, 4), ("height", result.height, 5),
        ("grid_start_row", GRID_START_ROW, 6), ("grid_start_col", GRID_START_COL, 7),
    ):
        data_ws.cell(row, 1, key)
        data_ws.cell(row, 2, value)
    data_ws["A9"] = "PALETTE_START"
    data_ws.append(["code", "name", "r", "g", "b"])
    for color in full_palette:
        data_ws.append([color.code, color.name, *color.rgb])
    original_row = 11 + len(full_palette)
    data_ws.cell(original_row, 1, "ORIGINAL_GRID_START")
    data_ws.cell(original_row, 2, original_row + 1)
    for y in range(result.height):
        for x in range(result.width):
            data_ws.cell(original_row + 1 + y, 1 + x, result.palette[int(result.indices[y, x])].code)
    data_ws.sheet_state = "hidden"
    wb.save(path)
    return path


def _theme_colors(wb) -> list[str]:
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return []
    try:
        root = ElementTree.fromstring(raw)
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        scheme = root.find(".//a:clrScheme", ns)
        colors: list[str] = []
        if scheme is None:
            return colors
        for node in list(scheme):
            child = next(iter(node), None)
            value = None
            if child is not None:
                value = child.attrib.get("lastClr") if child.tag.endswith("sysClr") else child.attrib.get("val")
            colors.append(value[-6:].upper() if value else "000000")
        return colors
    except Exception:
        return []


def _apply_tint(channel: int, tint: float) -> int:
    if tint < 0:
        return round(channel * (1.0 + tint))
    return round(channel * (1.0 - tint) + 255 * tint)


def _cell_fill_rgb(cell, themes: list[str]) -> tuple[int, int, int] | None:
    if cell.fill.fill_type != "solid":
        return None
    color = cell.fill.fgColor
    value: str | None = None
    if color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]
    elif color.type == "indexed" and color.indexed is not None and int(color.indexed) < len(COLOR_INDEX):
        value = COLOR_INDEX[int(color.indexed)][-6:]
    elif color.type == "theme" and color.theme is not None and int(color.theme) < len(themes):
        value = themes[int(color.theme)]
    if not value:
        return None
    try:
        rgb = tuple(int(value[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return None
    tint = float(color.tint or 0.0)
    if tint:
        rgb = tuple(max(0, min(255, _apply_tint(v, tint))) for v in rgb)
    return rgb


def _nearest_color(rgb: tuple[int, int, int], palette: list[BeadColor]) -> tuple[int, float]:
    target_lab = rgb_to_lab(np.asarray([rgb], dtype=np.uint8))
    palette_lab = rgb_to_lab(np.asarray([c.rgb for c in palette], dtype=np.uint8))
    distances = pairwise_delta_e(target_lab, palette_lab)[0]
    index = int(np.argmin(distances))
    return index, float(distances[index])


def import_pattern_xlsx(path: str | Path) -> tuple[PatternResult, ExcelImportReport]:
    """Read a MOSAIBeads workbook, detecting fill-only and code-only edits."""
    path = Path(path)
    wb = load_workbook(path, data_only=False)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError("这不是 MOSAIBeads 导出的 Excel：缺少隐藏的无损数据页。")
    data_ws = wb[DATA_SHEET]
    if data_ws["A1"].value != FORMAT_MARKER:
        raise ValueError("Excel 的 MOSAIBeads 格式标记无效。")
    version = int(data_ws["B2"].value or 0)
    if version > FORMAT_VERSION:
        raise ValueError(f"该 Excel 格式版本 {version} 高于当前软件支持的 {FORMAT_VERSION}。")
    brand = str(data_ws["B3"].value or "MARD 291")
    width, height = int(data_ws["B4"].value), int(data_ws["B5"].value)
    start_row, start_col = int(data_ws["B6"].value), int(data_ws["B7"].value)

    palette: list[BeadColor] = []
    row = 11
    while data_ws.cell(row, 1).value not in (None, "ORIGINAL_GRID_START"):
        code = str(data_ws.cell(row, 1).value)
        name = str(data_ws.cell(row, 2).value or code)
        rgb = tuple(int(data_ws.cell(row, col).value) for col in (3, 4, 5))
        palette.append(BeadColor(code, name, rgb))
        row += 1
    if data_ws.cell(row, 1).value != "ORIGINAL_GRID_START":
        raise ValueError("Excel 隐藏数据页缺少原始格子数据。")
    original_start = int(data_ws.cell(row, 2).value)
    original_codes = [
        [str(data_ws.cell(original_start + y, 1 + x).value or "").strip() for x in range(width)]
        for y in range(height)
    ]
    if not palette:
        raise ValueError("Excel 隐藏数据页的品牌色板为空。")

    ws = wb[PATTERN_SHEET] if PATTERN_SHEET in wb.sheetnames else wb.active
    themes = _theme_colors(wb)
    code_map = {c.code.strip().upper(): i for i, c in enumerate(palette)}
    indices = np.zeros((height, width), dtype=np.int32)
    report = ExcelImportReport(path, brand, width, height)
    nearest_cache: dict[tuple[int, int, int], tuple[int, float]] = {}

    for y in range(height):
        for x in range(width):
            cell = ws.cell(start_row + y, start_col + x)
            original_code = original_codes[y][x].upper()
            original_idx = code_map.get(original_code)
            if original_idx is None:
                raise ValueError(f"原始格子存在未知色号：{original_code}")
            text_code = str(cell.value or "").strip().upper()
            text_idx = code_map.get(text_code)
            fill_rgb = _cell_fill_rgb(cell, themes)
            original_rgb = palette[original_idx].rgb
            fill_changed = fill_rgb is not None and fill_rgb != original_rgb
            text_changed = bool(text_code) and text_code != original_code
            fill_idx, fill_delta = original_idx, 0.0
            if fill_changed and fill_rgb is not None:
                if fill_rgb not in nearest_cache:
                    nearest_cache[fill_rgb] = _nearest_color(fill_rgb, palette)
                fill_idx, fill_delta = nearest_cache[fill_rgb]

            if text_changed and text_idx is not None and not fill_changed:
                chosen = text_idx
                report.code_edits += 1
            elif fill_changed and not text_changed:
                chosen = fill_idx
                report.fill_edits += 1
            elif fill_changed and text_changed:
                report.fill_edits += 1
                report.code_edits += 1
                if text_idx is not None and palette[text_idx].rgb == fill_rgb:
                    chosen = text_idx
                else:
                    chosen = fill_idx
                    report.conflicts += 1
            elif text_idx is not None:
                chosen = text_idx
                report.unchanged += 1
            else:
                chosen = original_idx
                report.conflicts += 1
            if fill_changed and fill_delta > 0.5:
                report.approximate_matches += 1
            indices[y, x] = chosen

    rgb_grid = np.asarray([c.rgb for c in palette], dtype=np.uint8)[indices]
    result = PatternResult(
        indices=indices,
        palette=palette,
        selected_source_indices=np.arange(len(palette), dtype=np.int32),
        source_rgb=rgb_grid.copy(), sampled_rgb=rgb_grid.copy(),
        saliency=np.zeros((height, width), dtype=np.float32),
        profile="Excel 手工编辑",
        metadata={
            "brand_palette": brand, "bead_count": width * height,
            "excel_source": str(path), "excel_fill_edits": report.fill_edits,
            "excel_code_edits": report.code_edits,
            "excel_approximate_matches": report.approximate_matches,
            "excel_conflicts": report.conflicts,
        },
    )
    return result, report
