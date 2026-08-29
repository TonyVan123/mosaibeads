from __future__ import annotations

import os
from pathlib import Path
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import numpy as np
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .engine import PatternResult
from .excel_io import ExcelImportReport, export_pattern_xlsx, import_pattern_xlsx
from .exporter import export_bundle, render_clean
from .palettes import load_palette


class ExcelConverter:
    BG = "#101827"
    PANEL = "#1F2A3B"
    TEXT = "#F8FAFC"
    MUTED = "#AAB5C5"
    ACCENT = "#FF9F43"

    def __init__(self, root: tk.Tk, initial_path: str | None = None) -> None:
        self.root = root
        self.root.title("MOSAIBeads Excel 转拼豆图纸")
        self.root.geometry("1120x820")
        self.root.minsize(850, 650)
        self.root.configure(bg=self.BG)
        self.result: PatternResult | None = None
        self.report: ExcelImportReport | None = None
        self.preview_photo: ImageTk.PhotoImage | None = None
        self._style()
        self._build_ui()
        if initial_path:
            self.root.after(120, lambda: self.load_excel(initial_path))

    def _style(self) -> None:
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Dark.TFrame", background=self.BG)
        style.configure("Panel.TFrame", background=self.PANEL)
        style.configure("Title.TLabel", background=self.BG, foreground=self.TEXT, font=("Microsoft YaHei UI", 18, "bold"))
        style.configure("Dark.TLabel", background=self.BG, foreground=self.TEXT, font=("Microsoft YaHei UI", 10))
        style.configure("Panel.TLabel", background=self.PANEL, foreground=self.TEXT, font=("Microsoft YaHei UI", 10))
        style.configure("Accent.TButton", font=("Microsoft YaHei UI", 11, "bold"), padding=(14, 9))
        style.configure("Tool.TButton", font=("Microsoft YaHei UI", 10), padding=(11, 8))

    def _build_ui(self) -> None:
        header = ttk.Frame(self.root, style="Dark.TFrame", padding=(20, 15))
        header.pack(fill="x")
        ttk.Label(header, text="MOSAIBeads Excel → 拼豆图纸", style="Title.TLabel").pack(side="left")
        ttk.Button(header, text="选择 Excel", style="Tool.TButton", command=self.choose_excel).pack(side="right", padx=5)
        self.export_button = ttk.Button(header, text="导出完整图纸", style="Accent.TButton", command=self.export_result, state="disabled")
        self.export_button.pack(side="right", padx=5)

        note = tk.Label(
            self.root,
            text="支持 MOSAIBeads 导出的 .xlsx。只改背景色时按背景色转换；只改色号时按色号转换；任意 RGB 自动匹配最近的品牌豆色。",
            bg="#253247", fg=self.TEXT, padx=15, pady=10, anchor="w", font=("Microsoft YaHei UI", 10),
        )
        note.pack(fill="x", padx=20, pady=(0, 12))

        body = ttk.Frame(self.root, style="Dark.TFrame", padding=(20, 0, 20, 16))
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)
        self.preview = tk.Label(body, text="点击“选择 Excel”载入图纸", bg="#E8ECF1", fg="#4B5563", font=("Microsoft YaHei UI", 15))
        self.preview.grid(row=0, column=0, sticky="nsew", padx=(0, 14))

        panel = ttk.Frame(body, style="Panel.TFrame", padding=16, width=300)
        panel.grid(row=0, column=1, sticky="nsew")
        panel.grid_propagate(False)
        ttk.Label(panel, text="转换信息", style="Panel.TLabel", font=("Microsoft YaHei UI", 13, "bold")).pack(anchor="w", pady=(0, 12))
        self.path_var = tk.StringVar(value="尚未选择文件")
        self.summary_var = tk.StringVar(value="")
        self.detail_var = tk.StringVar(value="")
        tk.Label(panel, textvariable=self.path_var, bg=self.PANEL, fg=self.MUTED, wraplength=265, justify="left", anchor="w", font=("Microsoft YaHei UI", 9)).pack(fill="x", pady=(0, 15))
        tk.Label(panel, textvariable=self.summary_var, bg=self.PANEL, fg=self.TEXT, wraplength=265, justify="left", anchor="w", font=("Microsoft YaHei UI", 11, "bold")).pack(fill="x", pady=(0, 15))
        tk.Label(panel, textvariable=self.detail_var, bg=self.PANEL, fg=self.MUTED, wraplength=265, justify="left", anchor="w", font=("Microsoft YaHei UI", 9)).pack(fill="x")

        status = tk.Frame(self.root, bg="#0A111D", padx=15, pady=6)
        status.pack(fill="x")
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(status, textvariable=self.status_var, bg="#0A111D", fg=self.MUTED, anchor="w", font=("Microsoft YaHei UI", 9)).pack(fill="x")

    def choose_excel(self) -> None:
        path = filedialog.askopenfilename(title="选择可编辑拼豆 Excel", filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")], parent=self.root)
        if path:
            self.load_excel(path)

    def load_excel(self, path: str | Path) -> None:
        self.status_var.set("正在读取背景色、色号和隐藏的品牌色板……")
        self.root.update_idletasks()
        try:
            result, report = import_pattern_xlsx(path)
        except Exception as exc:
            messagebox.showerror("无法转换 Excel", str(exc), parent=self.root)
            self.status_var.set("载入失败。")
            return
        self.result, self.report = result, report
        self.path_var.set(str(Path(path)))
        self.summary_var.set(f"{report.width} × {report.height}\n{report.width * report.height} 颗 · {len(result.counts())} 种已用色\n品牌色板：{report.brand}")
        self.detail_var.set(
            f"识别到的 Excel 修改\n\n背景填充：{report.fill_edits} 格\n文字色号：{report.code_edits} 格\n近似颜色：{report.approximate_matches} 格\n填充/文字冲突：{report.conflicts} 格\n\n冲突时以背景填充色为准。"
        )
        self._render_preview()
        self.export_button.configure(state="normal")
        self.status_var.set("转换完成，可导出 PNG、PDF、材料表、JSON 和新的可编辑 Excel。")

    def _render_preview(self) -> None:
        if self.result is None:
            return
        image = render_clean(self.result, max(6, min(18, 780 // max(self.result.width, self.result.height))), bead_holes=True)
        image.thumbnail((760, 690), Image.Resampling.LANCZOS)
        self.preview_photo = ImageTk.PhotoImage(image)
        self.preview.configure(image=self.preview_photo, text="")

    def export_result(self) -> None:
        if self.result is None or self.report is None:
            return
        folder = filedialog.askdirectory(title="选择图纸导出文件夹", parent=self.root)
        if not folder:
            return
        source = self.report.source_path.stem
        out = Path(folder) / f"{source}_转换图纸"
        try:
            outputs = export_bundle(self.result, out, source)
            report_path = out / f"{source}_Excel转换报告.txt"
            report_path.write_text(self.report.message + "\n\n规则：背景色与文字冲突时以背景色为准；非品牌 RGB 使用 CIEDE2000 色差匹配最近品牌色。", encoding="utf-8")
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc), parent=self.root)
            return
        self.status_var.set(f"已导出 {len(outputs) + 1} 个文件：{out}")
        if messagebox.askyesno("导出完成", f"已导出到：\n{out}\n\n是否打开文件夹？", parent=self.root):
            os.startfile(out)


def run(initial_path: str | None = None) -> None:
    root = tk.Tk()
    ExcelConverter(root, initial_path)
    root.mainloop()


def smoke_test() -> None:
    palette = load_palette("MARD 291")[:6]
    indices = np.arange(20, dtype=np.int32).reshape(4, 5) % len(palette)
    rgb = np.asarray([c.rgb for c in palette], dtype=np.uint8)[indices]
    result = PatternResult(indices, palette, np.arange(len(palette)), rgb.copy(), rgb.copy(), np.zeros((4, 5), dtype=np.float32), "测试", {"brand_palette": "MARD 291"})
    with tempfile.TemporaryDirectory() as folder:
        path = export_pattern_xlsx(result, Path(folder) / "roundtrip.xlsx")
        wb = load_workbook(path)
        ws = wb["拼豆图纸"]
        for col in range(2, 7):
            ws.cell(3, col).fill = PatternFill("solid", fgColor="808080")
        wb.save(path)
        restored, report = import_pattern_xlsx(path)
        if restored.indices.shape != (4, 5) or report.fill_edits != 5:
            raise RuntimeError("Excel 整行填充色回读测试失败")
        if len(set(int(v) for v in restored.indices[1])) != 1:
            raise RuntimeError("Excel 整行灰色没有统一映射到品牌豆色")
